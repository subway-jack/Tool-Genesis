# -*- coding: utf-8 -*-
"""
ExcelToolkit
------------
Expose ONLY `extract_excel_content` as an agent tool.
Internal helpers (e.g., `_convert_to_markdown`, `write_excel`) are NOT tools.

Dependencies:
    - pandas
    - openpyxl
    - xls2xlsx (only when converting legacy .xls)
    - tabulate (for markdown table rendering)
"""

from __future__ import annotations

import os
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook

from src.core.toolkits import FunctionTool
from src.core.toolkits.base import BaseToolkit


class ExcelToolkit(BaseToolkit):
    r"""Toolkit to extract detailed cell information from Excel/CSV files.

    Only `extract_excel_content` is exposed to the agent tool system.
    This class does **not** process docx/pdf/pptx; the original docstring was misleading.

    Args:
        timeout (Optional[float]): Timeout (seconds) used by BaseToolkit's with_timeout wrapper.
    """

    def __init__(self, timeout: Optional[float] = None) -> None:
        super().__init__(timeout=timeout)

    # --------------------------- internal helpers (NOT tools) ---------------------------

    def _convert_to_markdown(self, df: pd.DataFrame) -> str:
        r"""Convert DataFrame to a Markdown table.

        Args:
            df (pd.DataFrame): DataFrame containing the sheet data.

        Returns:
            str: Markdown formatted table.
        """
        from tabulate import tabulate
        # Keep index off to avoid extra index column unless needed
        md_table = tabulate(df, headers="keys", tablefmt="pipe", showindex=False)
        return str(md_table)

    # ------------------------------ exported tool method -------------------------------

    def extract_excel_content(self, document_path: str) -> str:
        r"""Extract detailed cell information from an Excel/CSV file (all sheets).

        Supported formats: .xls, .xlsx, .csv
        - .xls will be converted to .xlsx via xls2xlsx.
        - .csv will be read as a single "sheet" view.

        For each sheet:
          - Collect (truncated) cell info: address (e.g., A1), value, font_color, fill_color
          - Provide a Markdown preview of the sheet (truncated)

        Args:
            document_path (str): Path to the Excel/CSV file.

        Returns:
            str: Formatted string including per-sheet cell metadata and markdown preview.
        """
        if not (
            document_path.endswith(".xls")
            or document_path.endswith(".xlsx")
            or document_path.endswith(".csv")
        ):
            return (
                f"Failed to process file {document_path}: "
                f"not an Excel/CSV format (.xls/.xlsx/.csv)."
            )

        # Handle CSV directly (single-sheet view)
        if document_path.endswith(".csv"):
            try:
                df = pd.read_csv(document_path)
                md_table = self._convert_to_markdown(df)
                return f"CSV File Processed:\n{md_table}"
            except Exception as e:
                return f"Failed to process file {document_path}: {e}"

        # Convert legacy .xls to .xlsx
        if document_path.endswith(".xls"):
            try:
                from xls2xlsx import XLS2XLSX  # lazy import
            except Exception:
                return (
                    "Failed to process .xls -> .xlsx: package `xls2xlsx` not found. "
                    "Install via `pip install xls2xlsx` or provide an .xlsx file."
                )
            try:
                output_path = document_path.replace(".xls", ".xlsx")
                x2x = XLS2XLSX(document_path)
                x2x.to_xlsx(output_path)
                document_path = output_path
            except Exception as e:
                return f"Failed to convert {os.path.basename(document_path)} to .xlsx: {e}"

        # Now we have an .xlsx path
        try:
            wb = load_workbook(document_path, data_only=True)
        except Exception as e:
            return f"Failed to open workbook {document_path}: {e}"

        sheet_info_list = []

        # Iterate through sheets and collect cell metadata
        for sheet in wb.sheetnames:
            try:
                ws = wb[sheet]
            except KeyError:
                # Defensive: skip if sheet not accessible
                continue

            cell_info_list = []
            # Iterate through used cell range
            for row in ws.iter_rows():
                for cell in row:
                    row_num = cell.row
                    col_letter = cell.column_letter
                    cell_value = cell.value

                    # Colors can be theme-based or rgb; we keep raw rgb if present
                    font_color = None
                    if cell.font and cell.font.color and "rgb=None" not in str(cell.font.color):
                        font_color = getattr(cell.font.color, "rgb", None)

                    fill_color = None
                    if cell.fill and getattr(cell.fill, "fgColor", None) and "rgb=None" not in str(cell.fill.fgColor):
                        fill_color = getattr(cell.fill.fgColor, "rgb", None)

                    # Excel address is like "A1" (ColumnLetter + RowNumber)
                    cell_info_list.append(
                        {
                            "index": f"{col_letter}{row_num}",
                            "value": cell_value,
                            "font_color": font_color,
                            "fill_color": fill_color,
                        }
                    )

            # For a human-friendly view, read the sheet into a DataFrame
            try:
                sheet_df = pd.read_excel(document_path, sheet_name=sheet, engine="openpyxl")
                markdown_content = self._convert_to_markdown(sheet_df)
            except Exception as e:
                markdown_content = f"(Failed to render sheet '{sheet}' to markdown: {e})"

            sheet_info_list.append(
                {
                    "sheet_name": sheet,
                    "cell_info_list": cell_info_list,
                    "markdown_content": markdown_content,
                }
            )

        # Truncate oversized outputs to keep the tool response manageable
        MAX_CHAR_LENGTH = 5000
        result_str_parts: List[str] = []
        for sheet_info in sheet_info_list:
            raw_cell_info = str(sheet_info["cell_info_list"])
            raw_md = str(sheet_info["markdown_content"])

            cell_info_display = raw_cell_info
            md_display = raw_md

            if len(raw_cell_info) > MAX_CHAR_LENGTH:
                cell_info_display = raw_cell_info[:MAX_CHAR_LENGTH]
                cell_info_display += f"... (Truncated, total length is {len(raw_cell_info)})"

            if len(raw_md) > MAX_CHAR_LENGTH:
                md_display = raw_md[:MAX_CHAR_LENGTH]
                md_display += (
                    f"... (Truncated, total length is {len(raw_md)}, "
                    f"please write python code to get the full content)"
                )

            result_str_parts.append(
                f"Sheet Name: {sheet_info['sheet_name']}\n"
                f"Cell information list:\n{cell_info_display}\n\n"
                f"Markdown View of the content:\n{md_display}\n"
                f"{'-'*40}"
            )

        return "\n".join(result_str_parts) if result_str_parts else "No sheets found."

    # ---------------------------- non-tool convenience ----------------------------

    def write_excel(self, file_path: str, content: list[dict]) -> None:
        """
        Write list of dictionaries to an Excel file.
        Not exported as a tool by default.

        Args:
            file_path (str): Path to the output Excel file.
            content (list[dict]): Data to write, where each dict is a row.
        """
        df = pd.DataFrame(content)
        df.to_excel(file_path, index=False)

    # ----------------------------- tool exposure control -----------------------------

    def get_tools(self) -> List[FunctionTool]:
        """
        Only expose `extract_excel_content` as a tool.
        If you later need `write_excel` as a tool, append it here explicitly.
        """
        return [FunctionTool(self.extract_excel_content)]